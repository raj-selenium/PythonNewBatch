from openpyxl import load_workbook

wb = load_workbook("C:/Users/rajka/PycharmProjects/Allure_PyTest_Se_POM_Framework_Final/resources/TestData.XLSX")
sheet = wb["Login_Data"]
#print(sheet.max_row)
#print(sheet.max_column)
#print(sheet.cell(3,1).value)

for row in range(1,sheet.max_row):
    for col in range(1,sheet.max_column):
        print(sheet.cell(row,col).value)

sheet.cell(5,2).value = "Y"

wb.save("C:/Users/rajka/PycharmProjects/Allure_PyTest_Se_POM_Framework_Final/resources/TestData.XLSX")
wb.close()